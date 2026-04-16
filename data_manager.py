import openpyxl
from openpyxl.styles import PatternFill
import os

class DataManager:
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.wb = None
        self.sheet = None
        
    def load(self):
        try:
            self.wb = openpyxl.load_workbook(self.excel_path)
            self.sheet = self.wb.active
            return True
        except Exception as e:
            print(f"Error loading Excel: {e}")
            return False

    def get_data(self, row_index):
        """
        获取指定行的数据
        row_index: 从 1 开始
        返回: (content, image_name)
        """
        if not self.sheet: return None, None
        
        # A列: 内容
        content = self.sheet.cell(row=row_index, column=1).value
        # B列: 图片名称
        image_name = self.sheet.cell(row=row_index, column=2).value
        
        return content, image_name

    def mark_failed(self, row_index):
        """
        标记指定行为红色
        """
        if not self.sheet: return
        
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        self.sheet.cell(row=row_index, column=1).fill = red_fill
        
        try:
            self.wb.save(self.excel_path)
        except Exception as e:
            print(f"Error saving Excel (failed mark): {e}")
